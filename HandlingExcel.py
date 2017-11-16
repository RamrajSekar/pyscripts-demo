import bs4 as bs
import urllib.request
import openpyxl as exl
import os
#To get the current working directory
print(os.getcwd())


##To create a new excel workbook
##wb = exl.Workbook()
##wb.get_sheet_names()
##sheet =wb.active
##sheet.title = 'spam spam'
##wb.save('example.xlsx')
##wb.get_sheet_names()


##print('<<<<<<<<<To open an existing excel file>>>>>>')
##wb = exl.load_workbook('example.xlsx')
###To get all the sheet names
##print(wb.get_sheet_names())
###To get a specific sheet name
##sheet = wb.get_sheet_by_name('mahathi')
##print(sheet)
##
###To get the sheet name without tag
##print(sheet.title)
##
####To get the sheet type
##print(type(sheet))
##a = sheet['A1']
##print(a.value)
###To print a particular cell value
##b = sheet['B1']
##print(b.value)
##
###To print a particular row value
##print('Row ' + str(b.row) + ', Column ' + b.column + ' is ' +b.value)
##
###To print a particular cell value
##print('Value in cell ' + b.coordinate + ' is ' +b.value)
##
###To get max rows in the sheet
##print('The row count in '+sheet.title+' is '+str(sheet.max_row))
##
###To get max cols in the sheet
##print('The column count in '+sheet.title+' is '+str(sheet.max_column))
##
##print('------To get all the row values from sheet----')
##for i in range(1,int(sheet.max_row)+1,1):
##    for j in range(1,int(sheet.max_column)+1,1):
##        print(i,sheet.cell(row=i, column=j).value)
##    print('-------End------ ')
##print('----To get all the values from a specific column----')
##for i in range(1,int(sheet.max_row)+1,1):
##    print(sheet.cell(row=i, column=2).value)
##
##print('----To get all cell names for which value exists----')
##print(tuple(sheet['A1':'C4']))
##
##print('----To get all value from each cells----')
##for rowOfCellobjs in sheet['A1':'C4']:
##    for cellObjs in rowOfCellobjs:
##        print(cellObjs.coordinate,cellObjs.value)
##    print('-------End------ ')spam spam
##
##print('....Another way to reading rows in spreadsheet.........')
##for row in range(1,sheet.max_row+1):
##    A1 = sheet['A' + str(row)].value
##    B1 = sheet['B' + str(row)].value
##    C1 = sheet['C' + str(row)].value
##    print(A1)
##    print(B1)
##    print(C1)


print('-------To create a new excel workbook-----')
wb = exl.Workbook()
wb.get_sheet_names()
sheet =wb.active
sheet.title = 'Python script write excel'
wb.save('WriteExcel.xlsx')

#To create sheet with default name
wb.create_sheet()
print(wb.get_sheet_names())

#To create sheet at a specified position with user defined name
wb.create_sheet(index=0,title='New Sheet')
wb.create_sheet(index=2,title='New Sheet1')
print(wb.get_sheet_names())

#To remove a sheet
wb.remove_sheet(wb.get_sheet_by_name('New Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('New Sheet1'))
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
print(wb.get_sheet_names())
wb.save('WriteExcel.xlsx')

#To write in cell
print('----To write values to cells----')
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)
wb.save('WriteExcel.xlsx')
