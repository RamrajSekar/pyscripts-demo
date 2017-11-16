import openpyxl as exl
import os
#import collections

print(os.getcwd())
wb = exl.load_workbook('example.xlsx')
print('----READING WORKSHEET----')

#To delete unwanted sheets
##for snames in wb.get_sheet_names():
##    sname = snames
##    #print(sname)
##    if sname != 'mahathi':
##        wb.remove_sheet(wb.get_sheet_by_name(sname))
##print(wb.get_sheet_names())
##wb.save('example.xlsx')

sheet = wb.get_active_sheet()
print('Active woorksheet is',sheet)
profit = 0
for row in range(2,sheet.max_row+1):
    Fruits = sheet.cell(row=row,column=1).value
    if Fruits is not None:
        Price = sheet.cell(row=row,column=2).value
        Kgs_sold = sheet.cell(row=row,column=3).value
        profit = Price * Kgs_sold
        #To Write in excel cell
        sheet.cell(row=row,column=4).value = profit
        print('The total sales of '+Fruits+ ' is $',profit)
wb.save('example.xlsx')
##x = []
##for row in range(2,sheet.max_row+1):
##    Fruits = sheet.cell(row=row,column=1).value
##    #Price = sheet.cell(row=row,column=2).value
##    x.append(Fruits)
##
###To find the duplicate values in list
##for i in range(0,len(x)):
##    for j in range(i+1, len(x)):
##        if x[i] == x[j]:
##            print('Duplicate',x[i])

#To get the duplicate count in list
##y = collections.Counter(x)
##print(y)
##print(list(y))
