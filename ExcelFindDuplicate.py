import openpyxl as pyxl
import os

wb = pyxl.load_workbook('example.xlsx')

sheet = wb.active

print(sheet)
#print(wb.get_sheet_names())
#print(sheet['A'+str(1)].value)

print('-------INCOMPLETE----------')
for i in range(1,sheet.max_row+1):
    #print(sheet.cell(row=2,column=1).value)
    fruits = sheet['A'+str(i)].value
    
    price_perKg = sheet['B'+str(i)].value
   
    sold_units = sheet['C'+str(i)].value

    if fruits > str(1):
        print('The duplicate fruit details are ',fruits)
    
    #print(fruits,' ',price_perKg,' ',sold_units,' ')
    
   
