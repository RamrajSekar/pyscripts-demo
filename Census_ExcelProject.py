import openpyxl as exl
import pprint
import os


print('Opening workbook.....!!!')
wb = exl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')

countyData = {}##Dictionary
print('Column count is: ', sheet.max_column)
print('Row count is: ',sheet.max_row)

##SOLUTION 1:
##l_census = []
##l_pop = []

##To get specific cell value
##print(sheet['C1'].value)

##for i in range(2,sheet.max_row+1):
##    county = sheet.cell(row=i,column=3).value
##    if county == 'Autauga':
##        census = sheet.cell(row=i,column=1).value
##        pop = sheet.cell(row=i,column=4).value
##        l_census.append(census)
##        l_pop.append(pop)
##print('The total census of Autauga is ',len(l_census))
##print('The total population of Autauga is ',sum(l_pop))
##

##SOLUTION 2:
print('Reading rows in spreadsheet.........')
for row in range(2,15,1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    #print('The population of state '+state+ ' and county ' +county+' is: ',pop)
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)
   
print(countyData)


## Open a new text file and write the contents of countyData to it.
##print('Writing results...')
##resultFile = open('census2010.py', 'w')
##resultFile.write('allData = ' + pprint.pformat(countyData))
##resultFile.close()
##print('Done.')
